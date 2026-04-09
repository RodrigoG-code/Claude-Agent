"""
render_excel.py
---------------
Generates a single Excel workbook summarising all brands for the month.

Sheets:
  1. "Overview"        — One row per brand, colour-coded flag cells, auto-filter
  2. "Recommendations" — One row per recommendation per brand (up to 5)
  3. "Raw Data"        — Pure numeric values, no formatting

Usage:
    python tools/render_excel.py --year 2026 --month 3

Inputs:
    .tmp/processed/peer_ranks_YYYY_MM.json

Outputs:
    .tmp/reports/excel/brand_performance_YYYY_MM.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
PROCESSED_DIR = BASE_DIR / ".tmp" / "processed"
EXCEL_DIR = BASE_DIR / ".tmp" / "reports" / "excel"

# Colours
COLOURS = {
    "Good":    {"fill": "D1FAE5", "font": "065F46"},
    "Average": {"fill": "FEF9C3", "font": "78350F"},
    "Poor":    {"fill": "FEE2E2", "font": "991B1B"},
    "Unknown": {"fill": "F3F4F6", "font": "6B7280"},
}
HEADER_FILL = PatternFill("solid", fgColor="1E1B4B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="EDE9FE")
SUBHEADER_FONT = Font(bold=True, color="1E1B4B", size=9)

THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

NAPPS_METRICS = [
    ("conversion_rate",       "Conv. Rate (%)",     "flag"),
    ("aov",                   "AOV (€)",            "flag"),
    ("add_to_cart_rate",      "Add-to-Cart (%)",    "flag"),
    ("cart_abandonment_rate", "Cart Abandon (%)",   "flag"),
    ("clv",                   "CLV (€)",            "flag"),
    ("churn",                 "Churn (%)",          "flag"),
    ("sessions",              "Sessions",           "flag"),
    ("refund_rate",           "Refund Rate (%)",    "flag"),
    ("repeat_purchase_rate",  "Repeat Purch. (%)",  "flag"),
    ("total_orders",          "Orders",             "flag"),
]

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


def flag_fill(flag: str) -> PatternFill:
    c = COLOURS.get(flag, COLOURS["Unknown"])
    return PatternFill("solid", fgColor=c["fill"])


def flag_font(flag: str, bold=False) -> Font:
    c = COLOURS.get(flag, COLOURS["Unknown"])
    return Font(color=c["font"], bold=bold, size=9)


def write_overview_sheet(ws, brands: list, year: int, month: int):
    ws.title = "Overview"

    headers = [
        "Brand ID", "Brand Name", "Sector",
        "Overall Flag", "Peer Rank (Overall)",
        "Total Sales (€)", "Sales MoM (%)",
    ]
    for metric_key, metric_label, _ in NAPPS_METRICS:
        headers.append(metric_label)
        headers.append(f"{metric_label} Flag")
        headers.append(f"{metric_label} MoM (%)")

    headers += ["Peer Rank (Sales)", "PDF URL", "Excel URL"]

    # Header row
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, brand in enumerate(brands, 2):
        overall_flag = brand.get("overall_flag", "Unknown")

        row = [
            brand.get("brand_id", ""),
            brand.get("brand_name", ""),
            brand.get("sector", ""),
            overall_flag,
            brand.get("peer_rank_overall", ""),
            brand.get("total_sales"),
            brand.get("sales_mom_pct"),
        ]

        for metric_key, _, _ in NAPPS_METRICS:
            row.append(brand.get(metric_key))
            row.append(brand.get(f"{metric_key}_flag", "Unknown"))
            row.append(brand.get(f"{metric_key}_mom_pct"))

        row += [
            brand.get("peer_rank_sales", ""),
            brand.get("pdf_url", ""),
            brand.get("excel_url", ""),
        ]

        ws.append(row)

        # Style the overall flag cell (column 4)
        flag_cell = ws.cell(row=row_idx, column=4)
        flag_cell.fill = flag_fill(overall_flag)
        flag_cell.font = flag_font(overall_flag, bold=True)
        flag_cell.alignment = Alignment(horizontal="center")

        # Style per-metric flag cells
        flag_col_offset = 7  # first flag col after overall (1-indexed: col 8 = conversion_rate flag)
        for i, (_, _, _) in enumerate(NAPPS_METRICS):
            flag_col = flag_col_offset + (i * 3) + 1 + 1  # value, flag, mom → flag is +2
            metric_flag = row[flag_col_offset + (i * 3) + 1]
            cell = ws.cell(row=row_idx, column=flag_col_offset + (i * 3) + 2)
            if isinstance(metric_flag, str) and metric_flag in COLOURS:
                cell.fill = flag_fill(metric_flag)
                cell.font = flag_font(metric_flag)
                cell.alignment = Alignment(horizontal="center")

        # Apply borders to all cells in row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Freeze top row
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    for col_idx in range(5, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def write_recommendations_sheet(ws, brands: list):
    ws.title = "Recommendations"

    headers = ["Brand Name", "Brand ID", "Sector", "Overall Flag", "Priority", "Recommendation", "Explanation"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    for brand in brands:
        insights = brand.get("insights", {})
        improvements = insights.get("top_improvements", []) if isinstance(insights, dict) else []

        for rec in improvements[:5]:
            priority = rec.get("priority", "")
            row = [
                brand.get("brand_name", ""),
                brand.get("brand_id", ""),
                brand.get("sector", ""),
                brand.get("overall_flag", ""),
                priority,
                rec.get("title", ""),
                rec.get("explanation", ""),
            ]
            ws.append(row)

            # Colour-code overall flag and priority
            row_idx = ws.max_row
            flag_cell = ws.cell(row=row_idx, column=4)
            flag = brand.get("overall_flag", "Unknown")
            flag_cell.fill = flag_fill(flag)
            flag_cell.font = flag_font(flag, bold=True)
            flag_cell.alignment = Alignment(horizontal="center")

            pri_cell = ws.cell(row=row_idx, column=5)
            priority_colours = {"High": "FEE2E2", "Medium": "FEF9C3", "Low": "DBEAFE"}
            priority_fonts   = {"High": "991B1B", "Medium": "78350F", "Low": "1E40AF"}
            if priority in priority_colours:
                pri_cell.fill = PatternFill("solid", fgColor=priority_colours[priority])
                pri_cell.font = Font(color=priority_fonts[priority], bold=True, size=9)
            pri_cell.alignment = Alignment(horizontal="center")

            # Wrap explanation text
            ws.cell(row=row_idx, column=7).alignment = Alignment(wrap_text=True)

            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 60


def write_raw_data_sheet(ws, brands: list):
    ws.title = "Raw Data"

    headers = [
        "snapshot_date", "brand_id", "brand_name", "sector",
        "total_sales", "sales_mom_pct",
        "conversion_rate", "conversion_rate_mom_pct",
        "aov", "aov_mom_pct",
        "add_to_cart_rate", "add_to_cart_mom_pct",
        "cart_abandonment_rate", "cart_abandonment_mom_pct",
        "clv", "clv_mom_pct",
        "churn", "churn_mom_pct",
        "sessions", "sessions_mom_pct",
        "refund_rate", "refund_rate_mom_pct",
        "repeat_purchase_rate", "repeat_purchase_rate_mom_pct",
        "total_orders", "total_orders_mom_pct",
        "overall_flag", "overall_score",
        "peer_rank_overall", "peer_rank_sales",
    ]

    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    for brand in brands:
        period = brand.get("period", "")
        snapshot_date = f"{period}-01" if period else ""
        row = [
            snapshot_date,
            brand.get("brand_id", ""),
            brand.get("brand_name", ""),
            brand.get("sector", ""),
            brand.get("total_sales"),
            brand.get("sales_mom_pct"),
            brand.get("conversion_rate"),
            brand.get("conversion_rate_mom_pct"),
            brand.get("aov"),
            brand.get("aov_mom_pct"),
            brand.get("add_to_cart_rate"),
            brand.get("add_to_cart_mom_pct"),
            brand.get("cart_abandonment_rate"),
            brand.get("cart_abandonment_mom_pct"),
            brand.get("clv"),
            brand.get("clv_mom_pct"),
            brand.get("churn"),
            brand.get("churn_mom_pct"),
            brand.get("sessions"),
            brand.get("sessions_mom_pct"),
            brand.get("refund_rate"),
            brand.get("refund_rate_mom_pct"),
            brand.get("repeat_purchase_rate"),
            brand.get("repeat_purchase_rate_mom_pct"),
            brand.get("total_orders"),
            brand.get("total_orders_mom_pct"),
            brand.get("overall_flag", ""),
            brand.get("overall_score"),
            brand.get("peer_rank_overall"),
            brand.get("peer_rank_sales"),
        ]
        ws.append(row)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).border = THIN_BORDER

    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def main():
    parser = argparse.ArgumentParser(description="Render Excel summary for all brands.")
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, required=True)
    args = parser.parse_args()

    EXCEL_DIR.mkdir(parents=True, exist_ok=True)

    input_path = PROCESSED_DIR / f"peer_ranks_{args.year}_{args.month:02d}.json"
    if not input_path.exists():
        print(f"ERROR: {input_path} not found. Run generate_insights.py first.")
        sys.exit(1)

    with open(input_path) as f:
        brands = json.load(f)

    # Sort by overall_score descending (Good brands first)
    brands.sort(key=lambda b: b.get("overall_score") or 0, reverse=True)

    print(f"Building Excel workbook for {len(brands)} brands...")

    wb = Workbook()

    # Remove default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_overview = wb.create_sheet("Overview")
    ws_recs = wb.create_sheet("Recommendations")
    ws_raw = wb.create_sheet("Raw Data")

    write_overview_sheet(ws_overview, brands, args.year, args.month)
    write_recommendations_sheet(ws_recs, brands)
    write_raw_data_sheet(ws_raw, brands)

    period_label = f"{MONTH_NAMES[args.month]} {args.year}"
    output_path = EXCEL_DIR / f"brand_performance_{args.year}_{args.month:02d}.xlsx"
    wb.save(str(output_path))

    size_kb = output_path.stat().st_size // 1024
    print(f"Excel saved: {output_path} ({size_kb} KB)")
    print(f"  Overview: {len(brands)} brands")
    print(f"  Recommendations: rows for up to {len(brands) * 5} entries")
    print(f"  Raw Data: {len(brands)} rows")


if __name__ == "__main__":
    main()
