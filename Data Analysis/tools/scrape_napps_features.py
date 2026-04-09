"""
Generate Excel file with all Napps platform features extracted from crawled data.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parent.parent
TMP = ROOT / ".tmp"

FEATURES = [
    ("Home Builder (Drag & Drop)", "Visual drag-and-drop editor to design your app's home screen."),
    ("PDP Builder (Product Detail Page)", "Builder for customizing product detail pages with multiple layout options."),
    ("Menu Builder", "Customizable navigation menu with drag-and-drop blocks."),
    ("NIA (AI Assistant)", "Integrated AI assistant that helps build more effective apps."),
    ("Multi-Tab Home", "Create multiple tabs on the home screen to segment content for different customer groups."),
    ("Schedule Home Designs", "Schedule new home screen designs to go live at specific times."),
    ("Home Design per Market", "Customize home page content for each market/country."),
    ("Menu Design per Market", "Highlight different collections for specific markets in the navigation menu."),
    ("PDP Design per Market", "Customize product detail page layouts with different metafields by market."),
    ("Exclusive Content", "Turn loyal customers into brand insiders with app-only content."),
    ("Exclusive Access (Store Locking)", "Transform the mobile app into a VIP experience."),
    ("App-Exclusive Products", "Create products available only through the mobile app."),
    ("App-Exclusive Discounts", "Offer discounts that are only redeemable within the mobile app."),
    ("Loyalty Program Integration", "Native integrations with loyalty platforms (Growave, Smile, LoyaltyLion, Yotpo, Stamped)."),
    ("Drops", "Create excitement and exclusivity with product drops."),
    ("Lookbook", "Inspire customers with curated lookbooks."),
    ("Unlimited Push Notifications", "Send unlimited custom push notifications to app users on all plans."),
    ("AI-Generated Notifications", "Use AI (NIA) to craft compelling and personalized notification copy."),
    ("Segmented Notifications", "Target notifications based on criteria such as country, interests, or user behaviour."),
    ("Scheduled Notifications", "Schedule notifications to send at the optimal time."),
    ("Cart Abandonment Notifications", "Automated push notifications that remind customers about items left in their cart."),
    ("Back in Stock Notifications", "Automatically notify customers when a product they were interested in is back in stock."),
    ("Wishlisted Discount Notifications", "Automatically notify customers when a product on their wishlist goes on sale."),
    ("Order Updated Notifications", "Keep customers informed with real-time order status updates."),
    ("Klaviyo Integration (Push)", "Boost Klaviyo flows with push notifications."),
    ("Ometria Integration (Push)", "Seamlessly integrate push notifications into Ometria's automated customer journeys."),
    ("Theme Options", "Complete control over every design element in the app."),
    ("Custom Fonts", "Bring your own brand fonts into the app."),
    ("Custom Icons", "Customize app icons to match your brand's aesthetic."),
    ("Cart Customization", "Customize the cart page with features that drive both conversion and AOV."),
    ("Customer Profile Options", "Create custom sections with information for customers."),
    ("Branding Options", "Extensive customization options to strengthen brand identity."),
    ("Localization / Multi-Language", "Full multi-language support."),
    ("Full RTL Support", "Complete right-to-left layout support for Arabic and Hebrew markets."),
    ("Multi-Currency", "Support for multiple currencies."),
    ("Shoppable Videos", "TikTok-style shoppable video feed within the app."),
    ("Product Tags / Badges", "Create custom product badges using Shopify metaobjects."),
    ("Quick Add to Cart", "Allow customers to add items to cart quickly without navigating to the full product page."),
    ("Low Stock Indicator", "Show low stock warnings on products to create urgency."),
    ("Gift Cards", "Support for digital gift cards."),
    ("Gift Wraps", "Offer gift wrapping as an option in the cart."),
    ("Free Gift", "Offer a free gift to customers who spend above a configurable threshold."),
    ("Wishlist", "Allow customers to save and track products they're interested in."),
    ("Product Search, Sorting & Filtering", "Fast, intuitive search with sorting and filtering options."),
    ("Product Metafields", "Display custom product information fields using Shopify metafields."),
    ("Metafield / Metaobject Support", "Bring all your Shopify metafield and metaobject content into the app."),
    ("Product Stories", "Add videos to product pages in an Instagram Stories-style experience block."),
    ("Smart App Banner", "Banner displayed on your mobile website promoting the app."),
    ("Desktop QR Code Banner", "QR code displayed on desktop website for app download."),
    ("Native Android & iOS App", "Fully native apps for both Android and iOS."),
    ("White Label App", "Fully branded app experience with no Napps branding."),
    ("Preview Mobile App", "Test and refine your app before launch."),
    ("ASO Support", "Bring your own App Store Optimization metadata."),
    ("App Listing Support", "Guidance and tools for creating compelling app store listings."),
    ("Real-Time Sync with Shopify", "All backoffice information synced in real-time with your Shopify store."),
    ("Napps Analytics Dashboard", "Built-in analytics with key metrics."),
    ("Firebase Analytics Integration", "Connect Firebase Analytics for detailed event tracking."),
    ("Facebook SDK Integration", "Integrate Facebook SDK for conversion tracking and retargeting."),
    ("MS Clarity Integration", "Integration with Microsoft Clarity for session recordings and heatmaps."),
    ("GDPR Ready", "Built-in GDPR compliance features."),
    ("Klaviyo", "Native integration for email/SMS marketing automation."),
    ("Growave", "Native loyalty, reviews, and wishlist integration."),
    ("Judge.me", "Native product review integration."),
    ("Stamped", "Native reviews and loyalty integration."),
    ("Smile.io", "Native loyalty program integration."),
    ("LoyaltyLion", "Native loyalty program integration for reward programs."),
    ("Yotpo", "Native reviews, loyalty, and UGC integration."),
    ("Wishlist Plus", "Native wishlist integration."),
    ("Meta (Facebook/Instagram)", "Native Meta integration for advertising and social commerce."),
    ("Ometria", "Native CRM and marketing automation integration."),
    ("ZeroPact", "Native integration for sustainability and carbon offset features."),
    ("Shopify New Customer Accounts", "Native support for Shopify's new customer account system."),
    ("Custom Integrations", "Tailor your app with custom integrations."),
    ("Dedicated Account Manager", "Personalized support with a dedicated account manager."),
    ("Implementation Specialist", "Guided setup process with implementation specialists."),
    ("Design Support", "Expert design support tailored to your brand's needs."),
    ("Email Support", "Dedicated email support across all plans."),
    ("Napps Help Center", "Comprehensive self-service knowledge base."),
    ("Constant Upgrades", "Regular platform updates with new features."),
    ("Community & Feature Requests", "Active user community for sharing experiences and requesting features."),
    ("App Usage Metrics", "Track how customers use your app."),
    ("Behavior Segmentation", "Identify patterns in browsing and shopping behavior."),
    ("Retention & Repurchase Data", "Measure how many customers return and how often they buy again."),
    ("Product Discount Notifications", "Automatic alerts when a product's price drops."),
    ("Automatic Offers in Cart", "Reward customers with gifts or discounts at certain cart value."),
    ("Dynamic Discounts & Rewards", "Apply personalized promotions based on cart size or behavior."),
    ("Free Shipping Progress Bar", "Show customers how close they are to free shipping."),
    ("Rankings & Badges", "Reward the most active customers with rankings and badges."),
    ("Members-Only Content", "Share content available only to app users."),
    ("Countdown + Push", "Combine countdown timers with push notifications."),
    ("Early Access", "Give app users priority to shop new products before anyone else."),
    ("Mobile-Only Drops", "Launch collections that are only available through the app."),
    ("Deep Links for Sharing", "Create smart links for social sharing and influencer posts."),
    ("Campaigns & Notifications by Market", "Launch promotions for specific countries or regions."),
    ("Catalog per Market", "Offer different product catalogs depending on the region."),
    ("Custom Menu Dividers & Icons", "Make navigation clearer with unique icons and separators."),
    ("Custom Menu Tags", "Add labels like 'New', 'Sale' to menu items."),
    ("Menu Promotional Components", "Add banners and promotions inside the menu."),
    ("Dynamic Menus", "Show different menus depending on the audience."),
    ("Discount Tags", "Highlight products with dynamic labels."),
    ("Installments Display", "Show payment plans directly in the app."),
    ("NIA Layout Suggestions", "AI-powered tips for the best layouts and components."),
    ("NIA Product Recommendations", "Smart product suggestions tailored to each customer."),
    ("In-Store Stock Display", "Display real-time product availability in physical stores."),
    ("QR Code for Loyalty Points", "Let customers earn loyalty points from in-store purchases."),
    ("QR Code for Returns & Exchanges", "Customers can start returns or exchanges via QR scan."),
    ("QR Code for Store Connection", "Link physical shopping with digital experience."),
    ("QR Codes to Install App", "Turn in-store visitors into app users."),
    ("Complete the Look (Cross-sell)", "Display products that go well together on the product page."),
    ("Cross-Selling Components on PDP", "Suggest related products on the product page."),
    ("Multiple Product Layouts", "Create and manage several product page versions."),
    ("In-Store Stock on PDP", "Show real-time product availability on the product detail page."),
    ("Personalized Recommendations (NIA on PDP)", "AI-powered product suggestions on the product page."),
    ("Address & Preferences Management", "Let customers save multiple addresses and personal details."),
    ("Loyalty Status & Points Display", "Show loyalty level and point balance in profile."),
    ("Personalized Offers in Profile", "Display exclusive promotions based on customer profile."),
    ("Anonymous User Push", "Reach users who haven't logged in with push notifications."),
    ("Deep Links in Notifications", "Take customers directly to specific app areas from notifications."),
    ("Advanced Push Segmentation", "Target customers based on profile, behavior, or market."),
    ("Gamification", "Add challenges, levels, and badges to the shopping journey."),
    ("Interactive Quizzes", "Offer short quizzes to guide customers toward the right products."),
    ("Points System", "Reward customers with points for every purchase or action."),
    ("Global-e", "Enable cross-border selling with localized pricing."),
    ("Google Analytics 4", "Connect GA4 to track app usage and sales performance."),
    ("TikTok-Style Community Feed", "Build a dynamic content feed within the app."),
    ("Blog Component", "Add a block on the home page for blog articles."),
    ("Shop the Look", "Curated outfit or product combination modules."),
    ("Collection View Layouts", "Let users decide how they browse products."),
    ("Quick Access / Filter Bar", "A fixed navigation bar for faster access to search and filter."),
    ("Last Searches", "Save and display recent searches."),
    ("Color Swatches", "Display product color variants as visual swatches."),
    ("Social Logins (Facebook & Google)", "One-tap sign-in using Facebook or Google accounts."),
    ("Family Members Feature", "Parents can create and manage profiles for each child."),
    ("Connectif", "Native integration for marketing automation."),
    ("Doofinder", "Native search integration with accurate, personalized results."),
    ("Live Me Up", "Bring live shopping to your app."),
    ("Omniwallet", "Offer flexible payment options and digital wallets."),
    ("Rebuy", "AI-powered product recommendations, upsells, and cross-sells."),
    ("Recharge", "Native subscription management."),
    ("Reveni", "Instant refunds and faster returns integration."),
    ("Kiwi Sizing", "Size guide integration for accurate sizing information."),
    ("Growave Reviews", "Collect and display customer reviews through Growave."),
    ("Growave Wishlist", "Native Growave wishlist integration."),
    ("Searchanise", "Advanced search integration for accurate search results."),
    ("Shopify Bundles", "Integrate Shopify Bundles for pre-configured product bundles."),
    ("Shopify Checkout Blocks", "All Shopify checkout blocks integrated within the app."),
    ("Shopify Flows", "Add mobile app actions to existing Shopify Flows."),
    ("Shopify B2B", "Integrate B2B functionality into your mobile app."),
    ("Shopify Markets", "Full integration with Shopify Markets for cross-border commerce."),
    ("Shopify Translate & Adapt", "Integration with Shopify's translation tools."),
    ("Account & Order History", "Customers keep the same account as the Shopify website."),
    ("Shopify Checkout Integration", "Uses Shopify's native checkout for secure purchasing."),
    ("Design Blocks per Home", "Configurable number of design blocks per home screen."),
]


def build_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Napps Features"

    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D4D4D4"),
        right=Side(style="thin", color="D4D4D4"),
        top=Side(style="thin", color="D4D4D4"),
        bottom=Side(style="thin", color="D4D4D4"),
    )
    alt_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")

    for col, header in enumerate(["Feature", "Description"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for i, (feature, description) in enumerate(FEATURES, start=2):
        feat_cell = ws.cell(row=i, column=1, value=feature)
        desc_cell = ws.cell(row=i, column=2, value=description)
        feat_cell.font = Font(name="Calibri", size=11, bold=True)
        desc_cell.font = cell_font
        feat_cell.alignment = cell_alignment
        desc_cell.alignment = cell_alignment
        feat_cell.border = thin_border
        desc_cell.border = thin_border
        if i % 2 == 0:
            feat_cell.fill = alt_fill
            desc_cell.fill = alt_fill

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{len(FEATURES) + 1}"

    output_path = ROOT / "config" / "napps_features.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel saved to {output_path}")
    print(f"Total features: {len(FEATURES)}")
    return output_path


if __name__ == "__main__":
    build_excel()
