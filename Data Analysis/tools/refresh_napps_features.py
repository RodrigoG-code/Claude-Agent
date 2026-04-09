"""
refresh_napps_features.py
--------------------------
Automated Step 0 of the monthly pipeline.
1. Re-scrapes napps.io for all features
2. Compares against existing feature catalog
3. If new features found, uses Claude to map them to metrics
4. Updates config/napps_features.xlsx and config/napps_feature_context.md

Usage:
    python tools/refresh_napps_features.py
"""

import json
import os
import sys
import time
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()

BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_DIR = BASE_DIR / "config"
TMP_DIR = BASE_DIR / ".tmp"
TMP_DIR.mkdir(exist_ok=True)

FIRECRAWL_KEY = os.getenv("FIRECRAWL_API_KEY")
ANTHROPIC_KEY = os.getenv("ANTHROPIC_API_KEY")

SCRAPE_URLS = [
    "https://napps.io/homebuilder",
    "https://napps.io/productbuilder",
    "https://napps.io/builder",
    "https://napps.io/brand",
    "https://napps.io/automated",
    "https://napps.io/capabilities",
    "https://napps.io/integrations",
    "https://napps.io/faq",
    "https://napps.io/pricing",
    "https://napps.io/notifications",
    "https://napps.io/casestudy",
]


def step1_scrape():
    """Scrape napps.io and return all page markdown."""
    from firecrawl import FirecrawlApp

    if not FIRECRAWL_KEY:
        print("ERROR: FIRECRAWL_API_KEY not set in .env")
        sys.exit(1)

    app = FirecrawlApp(api_key=FIRECRAWL_KEY)
    all_markdown = []

    print("Crawling napps.io...")
    try:
        result = app.crawl("https://napps.io/", limit=50, scrape_options={"formats": ["markdown"]}, poll_interval=5)
        pages = result.data if hasattr(result, 'data') else result.get("data", [])
        for page in pages:
            md = page.markdown if hasattr(page, 'markdown') else page.get("markdown", "")
            if md:
                all_markdown.append(md)
        print(f"  Crawled {len(pages)} pages")
    except Exception as e:
        print(f"  Crawl failed: {e}")

    print("Scraping feature pages...")
    for url in SCRAPE_URLS:
        try:
            result = app.scrape(url, formats=["markdown"])
            md = result.markdown if hasattr(result, 'markdown') else result.get("markdown", "")
            if md:
                all_markdown.append(md)
            print(f"  OK: {url}")
        except Exception as e:
            print(f"  FAILED: {url} — {e}")
        time.sleep(0.3)

    combined = "\n\n===PAGE BREAK===\n\n".join(all_markdown)
    (TMP_DIR / "napps_refresh_all.txt").write_text(combined, encoding="utf-8")
    print(f"Total content: {len(combined)} chars from {len(all_markdown)} pages")
    return combined


def step2_extract_features(scraped_content):
    """Use Claude to extract all features from scraped content."""
    import anthropic

    if not ANTHROPIC_KEY:
        print("ERROR: ANTHROPIC_API_KEY not set. Skipping feature extraction.")
        return None

    existing_context = (CONFIG_DIR / "napps_feature_context.md").read_text(encoding="utf-8")
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

    if len(scraped_content) > 150000:
        scraped_content = scraped_content[:150000]

    print("Asking Claude to identify new features...")
    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": f"""I have two inputs:

1. EXISTING FEATURE CONTEXT (the current reference document):
{existing_context[:30000]}

2. FRESHLY SCRAPED CONTENT from napps.io:
{scraped_content[:100000]}

Compare the scraped content against the existing feature context. Identify any NEW features that are mentioned in the scraped content but NOT already documented.

For each new feature found, respond with valid JSON (no markdown wrapping):
{{
  "new_features_found": true/false,
  "features": [
    {{
      "name": "Feature Name",
      "description": "1-2 sentence description.",
      "relevant_metrics": ["conversion_rate", "cart_abandonment_rate"],
      "mechanism": "How this feature moves those metrics.",
      "section_to_add_to": "CONVERSION RATE"
    }}
  ],
  "summary": "Brief summary of what changed."
}}

If no new features are found, return:
{{
  "new_features_found": false,
  "features": [],
  "summary": "No new features detected."
}}"""
        }],
    )

    raw_text = message.content[0].text.strip()
    if raw_text.startswith("```"):
        raw_text = raw_text.split("```")[1]
        if raw_text.startswith("json"):
            raw_text = raw_text[4:]
        raw_text = raw_text.strip()

    return json.loads(raw_text)


def step3_update_files(extraction_result):
    """Update the feature context and Excel if new features found."""
    if not extraction_result or not extraction_result.get("new_features_found"):
        print("No new features found. Files are up to date.")
        return False

    features = extraction_result.get("features", [])
    print(f"Found {len(features)} new feature(s):")
    for f in features:
        print(f"  + {f['name']}")

    context_path = CONFIG_DIR / "napps_feature_context.md"
    context = context_path.read_text(encoding="utf-8")

    new_section = "\n\n---\n\n## NEW FEATURES (Auto-detected)\n*Added by automated refresh — review and move to appropriate metric sections above.*\n\n"
    for f in features:
        metrics = ", ".join(f.get("relevant_metrics", []))
        new_section += f"- **{f['name']}** — {f['description']} *Relevant to: {metrics}. Mechanism: {f.get('mechanism', 'N/A')}*\n"

    if "## GENERAL CONTEXT FOR RECOMMENDATIONS" in context:
        context = context.replace(
            "## GENERAL CONTEXT FOR RECOMMENDATIONS",
            new_section + "\n## GENERAL CONTEXT FOR RECOMMENDATIONS"
        )
    else:
        context += new_section

    context_path.write_text(context, encoding="utf-8")
    print(f"Updated {context_path.name}")

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        excel_path = CONFIG_DIR / "napps_features.xlsx"
        wb = load_workbook(excel_path)
        ws = wb.active
        cell_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D4D4D4"),
            right=Side(style="thin", color="D4D4D4"),
            top=Side(style="thin", color="D4D4D4"),
            bottom=Side(style="thin", color="D4D4D4"),
        )
        new_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

        next_row = ws.max_row + 1
        for f in features:
            feat_cell = ws.cell(row=next_row, column=1, value=f["name"])
            desc_cell = ws.cell(row=next_row, column=2, value=f["description"])
            feat_cell.font = bold_font
            desc_cell.font = cell_font
            feat_cell.alignment = cell_alignment
            desc_cell.alignment = cell_alignment
            feat_cell.border = thin_border
            desc_cell.border = thin_border
            feat_cell.fill = new_fill
            desc_cell.fill = new_fill
            next_row += 1

        wb.save(excel_path)
        print(f"Updated {excel_path.name} ({ws.max_row - 1} total features)")
    except Exception as e:
        print(f"Excel update failed: {e}")

    log_path = TMP_DIR / "feature_refresh_log.json"
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(extraction_result, f, indent=2, ensure_ascii=False)
    print(f"Refresh log saved to {log_path}")

    return True


def main():
    print("=" * 60)
    print("NAPPS Feature Catalog Refresh")
    print("=" * 60)

    print("\n[1/3] Scraping napps.io...")
    scraped = step1_scrape()

    print("\n[2/3] Analyzing for new features...")
    try:
        result = step2_extract_features(scraped)
    except Exception as e:
        print(f"Feature extraction failed: {e}")
        print("Scraped content saved to .tmp/napps_refresh_all.txt for manual review.")
        return

    print("\n[3/3] Updating feature files...")
    updated = step3_update_files(result)

    print("\n" + "=" * 60)
    if updated:
        print("Feature catalog UPDATED with new features.")
        print("Review config/napps_feature_context.md — new features are in")
        print("the 'NEW FEATURES (Auto-detected)' section at the bottom.")
    else:
        print("Feature catalog is up to date. No changes needed.")
    print("=" * 60)


if __name__ == "__main__":
    main()
